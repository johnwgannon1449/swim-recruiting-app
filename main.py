import os, json, re, time
from flask import Flask, request, jsonify, send_from_directory
from dotenv import load_dotenv
import openpyxl

load_dotenv()
app = Flask(__name__, static_folder='static', static_url_path='')

# ---------------------------------------------------------------------------
# TEAM NAME NORMALIZATIONS
# Workbook cell widths truncate long names and add trailing qualifiers.
# These are the ONLY 9 places where workbook text differs from canonical names.
# Flagged explicitly per guardrail — no silent guessing.
# ---------------------------------------------------------------------------
TEAM_NAME_MAP = {
    "McDaniel College Swim Team":     ("McDaniel College",                 "trailing qualifier"),
    "Rochester Institute of Technol": ("Rochester Institute of Technology", "truncated at 30 chars"),
    "Rensselaer Polytechnic Institu": ("Rensselaer Polytechnic Institute",  "truncated at 30 chars"),
    "Union College (New York)":        ("Union College",                     "parenthetical qualifier"),
    "Massachusetts Institute of Tec": ("MIT",                               "truncated — schema uses 'MIT'"),
    "Worcester Polytechnic Institut": ("Worcester Polytechnic Institute",   "truncated at 30 chars"),
    "Wheaton College (Ma)":            ("Wheaton College (MA)",              "wrong casing"),
    "Whitworth University Swim Team": ("Whitworth University",              "trailing qualifier"),
    "California Institute of Techno": ("Caltech",                           "truncated — schema uses 'Caltech'"),
    "Saint Johns University":          ("Saint John's University",           "missing apostrophe"),
}

# ---------------------------------------------------------------------------
# JAMES — hardcoded swimmer profile (source of truth for all scoring runs)
# ---------------------------------------------------------------------------
JAMES = {
    "name":             "James",
    "gpa":              4.0,
    "sat":              1460,
    "satProjected":     1500,
    "mathSat":          720,
    "mathSatProjected": 760,
    "times": {
        "1650 Free":              "16:06",
        "1000 Free":              "9:30",
        "500 Free":               "4:37",
        "200 Free":               "1:43",
        "400 IM":                 "4:09",
        "200 IM":                 "1:56",
        "100 Breast":             "59.5",
        "50 Breast (Relay Split)": "25.68",
    },
    "vibe": {
        "campus":   "Small and tight-knit — everyone knows everyone",
        "friday":   "Library with 2–3 close friends",
        "academic": "Genuinely want to be well-rounded",
        "compete":  "Love pushing myself inside a team environment",
        "location": None,
        "career":   None,
    },
}

# ---------------------------------------------------------------------------
# SCHOOL_META — per-school metadata for all 76 programs
# Fields: accept (int %), satMedian (int), hiddenIvy (bool), stem (bool),
#         merit ("none"|"moderate"|"high"), location (str), vibe (str),
#         moonshot (bool, optional)
# Keys must match canonical names after TEAM_NAME_MAP normalization.
# ---------------------------------------------------------------------------
SCHOOL_META = {
    # ── CENTENNIAL ───────────────────────────────────────────────────────────
    "Johns Hopkins University": {
        "accept": 7, "satMedian": 1510, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Baltimore, MD",
        "vibe": "Research powerhouse where pre-med and STEM culture run the campus",
    },
    "Gettysburg College": {
        "accept": 43, "satMedian": 1230, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Gettysburg, PA",
        "vibe": "Historic campus with strong Greek life and leadership culture",
    },
    "Swarthmore College": {
        "accept": 7, "satMedian": 1505, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Swarthmore, PA",
        "vibe": "Academically intense and collaborative with Quaker roots",
    },
    "Dickinson College": {
        "accept": 48, "satMedian": 1235, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Carlisle, PA",
        "vibe": "Sustainability-focused with strong international programs",
    },
    "Franklin & Marshall College": {
        "accept": 34, "satMedian": 1280, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Lancaster, PA",
        "vibe": "Pre-professional culture with strong pre-law and alumni network",
    },
    "Ursinus College": {
        "accept": 67, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Collegeville, PA",
        "vibe": "Warm undergraduate-focused campus with strong research access",
    },
    "Washington College": {
        "accept": 75, "satMedian": 1095, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Chestertown, MD",
        "vibe": "Small waterfront campus; close community and strong writing tradition",
    },
    "McDaniel College": {
        "accept": 82, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Westminster, MD",
        "vibe": "Friendly and personal campus with strong teacher education programs",
    },
    # ── LIBERTY LEAGUE ───────────────────────────────────────────────────────
    "Ithaca College": {
        "accept": 68, "satMedian": 1180, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Ithaca, NY",
        "vibe": "Creative and artsy; media, music, and performance everywhere",
    },
    "Rochester Institute of Technology": {
        "accept": 73, "satMedian": 1295, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Rochester, NY",
        "vibe": "Tech and project-driven culture built around co-ops and real careers",
    },
    "Rensselaer Polytechnic Institute": {
        "accept": 63, "satMedian": 1410, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Troy, NY",
        "vibe": "Pure engineering culture; hard-working students who live for problems",
    },
    "Clarkson University": {
        "accept": 80, "satMedian": 1205, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Potsdam, NY",
        "vibe": "Tight-knit STEM community in the North Country; outdoorsy and close",
    },
    "Union College": {
        "accept": 38, "satMedian": 1335, "hiddenIvy": False, "stem": True,
        "merit": "moderate", "location": "Schenectady, NY",
        "vibe": "Liberal arts meets engineering; theme houses and strong traditions",
    },
    "Skidmore College": {
        "accept": 29, "satMedian": 1300, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Saratoga Springs, NY",
        "vibe": "Creative and arts-forward campus in a lively upstate NY town",
    },
    "Vassar College": {
        "accept": 18, "satMedian": 1455, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Poughkeepsie, NY",
        "vibe": "Progressive intellectual culture; students who love big ideas",
    },
    "St. Lawrence University": {
        "accept": 60, "satMedian": 1230, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Canton, NY",
        "vibe": "Outdoorsy North Country campus with a close community and athletics",
    },
    "Hobart and William Smith": {
        "accept": 58, "satMedian": 1185, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Geneva, NY",
        "vibe": "Lakeside dual-college campus with strong social scene and sailing",
    },
    "Bard College": {
        "accept": 64, "satMedian": 1265, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Annandale-on-Hudson, NY",
        "vibe": "Artsy and progressive with discussion-heavy classes and bohemian feel",
    },
    # ── NCAC ─────────────────────────────────────────────────────────────────
    "Denison University": {
        "accept": 28, "satMedian": 1325, "hiddenIvy": True, "stem": False,
        "merit": "high", "location": "Granville, OH",
        "vibe": "Beautiful hilltop campus; ambitious academics and a strong social scene",
    },
    "Kenyon College": {
        "accept": 33, "satMedian": 1370, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Gambier, OH",
        "vibe": "Literary and deeply intellectual; famous writing program in rural Ohio",
    },
    "John Carroll University": {
        "accept": 80, "satMedian": 1135, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "University Heights, OH",
        "vibe": "Jesuit values; service-oriented with strong business programs",
    },
    "DePauw University": {
        "accept": 67, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Greencastle, IN",
        "vibe": "Greek life-heavy with strong communications and music programs",
    },
    "Wabash College": {
        "accept": 69, "satMedian": 1165, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Crawfordsville, IN",
        "vibe": "All-male liberal arts with intense brotherhood and strong traditions",
    },
    "College of Wooster": {
        "accept": 57, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Wooster, OH",
        "vibe": "Every senior writes a thesis; close-knit with strong independent study",
    },
    "Oberlin College": {
        "accept": 33, "satMedian": 1385, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Oberlin, OH",
        "vibe": "Progressive and artistic; famous conservatory and engaged campus politics",
    },
    "Ohio Wesleyan University": {
        "accept": 83, "satMedian": 1150, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Delaware, OH",
        "vibe": "Emphasis on global experience; internship-focused with civic engagement",
    },
    "Wittenberg University": {
        "accept": 86, "satMedian": 1125, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Springfield, OH",
        "vibe": "Lutheran-rooted campus; personal and strong in teacher education",
    },
    # ── NESCAC ───────────────────────────────────────────────────────────────
    "Williams College": {
        "accept": 9, "satMedian": 1510, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Williamstown, MA",
        "vibe": "Consistently ranked #1 LAC; mountain campus with elite academics",
    },
    "Tufts University": {
        "accept": 9, "satMedian": 1500, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Medford, MA",
        "vibe": "Globally minded near Boston; research-intensive with elite academics",
    },
    "Amherst College": {
        "accept": 9, "satMedian": 1515, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Amherst, MA",
        "vibe": "Open curriculum, no required courses; fiercely intellectual with 5-College access",
    },
    "Connecticut College": {
        "accept": 38, "satMedian": 1315, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "New London, CT",
        "vibe": "Student self-governance model; students run nearly everything on campus",
    },
    "Bates College": {
        "accept": 13, "satMedian": 1430, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Lewiston, ME",
        "vibe": "Politically engaged and outdoorsy; tight community in coastal Maine",
    },
    "Hamilton College": {
        "accept": 14, "satMedian": 1440, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Clinton, NY",
        "vibe": "Writing-intensive; every major leads to a thesis on a beautiful rural campus",
    },
    "Bowdoin College": {
        "accept": 9, "satMedian": 1495, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Brunswick, ME",
        "vibe": "Outdoorsy and intellectual in coastal Maine; sustainability and community",
    },
    "Middlebury College": {
        "accept": 13, "satMedian": 1445, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Middlebury, VT",
        "vibe": "Environmental passion meets rigorous academics in a beautiful Vermont setting",
    },
    "Colby College": {
        "accept": 11, "satMedian": 1435, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Waterville, ME",
        "vibe": "Liberal arts in the Maine wilderness; entrepreneurial with a tight community",
    },
    "Trinity College": {
        "accept": 34, "satMedian": 1310, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Hartford, CT",
        "vibe": "Classic New England campus with strong city partnerships and Greek life",
    },
    "Wesleyan University": {
        "accept": 17, "satMedian": 1455, "hiddenIvy": True, "stem": False,
        "merit": "none", "location": "Middletown, CT",
        "vibe": "Quirky and politically active; film and social sciences define the culture",
    },
    # ── NEWMAC ───────────────────────────────────────────────────────────────
    "MIT": {
        "accept": 4, "satMedian": 1565, "hiddenIvy": False, "stem": True,
        "merit": "none", "moonshot": True, "location": "Cambridge, MA",
        "vibe": "The world's most famous STEM institution; unmatched resources and intensity",
    },
    "U.S. Coast Guard Academy": {
        "accept": 14, "satMedian": 1265, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "New London, CT",
        "vibe": "Military service academy; full scholarship, intense discipline, meaningful mission",
    },
    "Worcester Polytechnic Institute": {
        "accept": 58, "satMedian": 1370, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Worcester, MA",
        "vibe": "Project-based learning at a tech school with strong industry connections",
    },
    "Babson College": {
        "accept": 24, "satMedian": 1330, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Wellesley, MA",
        "vibe": "#1 entrepreneurship school; every freshman runs a real business for credit",
    },
    "Wheaton College (MA)": {
        "accept": 71, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Norton, MA",
        "vibe": "Small and personal campus reinventing itself with a bold connected curriculum",
    },
    "Springfield College": {
        "accept": 77, "satMedian": 1095, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Springfield, MA",
        "vibe": "Birthplace of basketball; health sciences and PT define the campus culture",
    },
    "Clark University": {
        "accept": 52, "satMedian": 1225, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Worcester, MA",
        "vibe": "Free fifth year for any grad program; research-first culture in an urban campus",
    },
    # ── NWC ──────────────────────────────────────────────────────────────────
    "Whitworth University": {
        "accept": 89, "satMedian": 1175, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Spokane, WA",
        "vibe": "Christian liberal arts in the Pacific Northwest with strong education programs",
    },
    "University of Puget Sound": {
        "accept": 87, "satMedian": 1200, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Tacoma, WA",
        "vibe": "NW outdoor culture meets classic liberal arts; tight-knit campus",
    },
    "Linfield University": {
        "accept": 88, "satMedian": 1120, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "McMinnville, OR",
        "vibe": "Small Oregon liberal arts in wine country; personal and community-focused",
    },
    "Whitman College": {
        "accept": 46, "satMedian": 1295, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Walla Walla, WA",
        "vibe": "Quirky Pacific NW intellectual gem with outdoor access and high grad school rates",
    },
    "Pacific Lutheran University": {
        "accept": 87, "satMedian": 1125, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Tacoma, WA",
        "vibe": "Lutheran values in the Pacific Northwest; strong music and education programs",
    },
    "Lewis & Clark College": {
        "accept": 65, "satMedian": 1275, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Portland, OR",
        "vibe": "Hippie-intellectual Portland culture; environmental law and social justice focus",
    },
    "Willamette University": {
        "accept": 81, "satMedian": 1190, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Salem, OR",
        "vibe": "NW LAC with strong law school connections and active civic engagement",
    },
    "George Fox University": {
        "accept": 93, "satMedian": 1090, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Newberg, OR",
        "vibe": "Christian university with strong community, athletics, and business programs",
    },
    # ── SCIAC ────────────────────────────────────────────────────────────────
    "Pomona-Pitzer": {
        "accept": 7, "satMedian": 1510, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Claremont, CA",
        "vibe": "Elite SoCal LAC in the Claremont Consortium; 5 colleges sharing resources",
    },
    "Claremont-Mudd-Scripps": {
        "accept": 9, "satMedian": 1490, "hiddenIvy": True, "stem": True,
        "merit": "none", "location": "Claremont, CA",
        "vibe": "Harvey Mudd's STEM intensity meets Scripps' creative and humanistic edge",
    },
    "Chapman University": {
        "accept": 52, "satMedian": 1230, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Orange, CA",
        "vibe": "Film school prestige meets SoCal sunshine; entrepreneurial and media-forward",
    },
    "Caltech": {
        "accept": 3, "satMedian": 1560, "hiddenIvy": False, "stem": True,
        "merit": "none", "moonshot": True, "location": "Pasadena, CA",
        "vibe": "Hardest STEM school to enter in America; Nobel laureates teach undergrads",
    },
    "Whittier College": {
        "accept": 73, "satMedian": 1100, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Whittier, CA",
        "vibe": "Liberal arts with strong Latino heritage and close community feel in SoCal",
    },
    "Cal Lutheran University": {
        "accept": 61, "satMedian": 1125, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Thousand Oaks, CA",
        "vibe": "Lutheran roots in Ventura County; strong business and communications programs",
    },
    "Occidental College": {
        "accept": 37, "satMedian": 1295, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Los Angeles, CA",
        "vibe": "Progressive urban LAC in Eagle Rock; politics and international relations culture",
    },
    "University of Redlands": {
        "accept": 67, "satMedian": 1140, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Redlands, CA",
        "vibe": "Inland Empire LAC; strong music, environmental studies, and pre-law culture",
    },
    "University of La Verne": {
        "accept": 65, "satMedian": 1080, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "La Verne, CA",
        "vibe": "Close-knit SoCal campus with strong business and criminal justice programs",
    },
    # ── UAA ──────────────────────────────────────────────────────────────────
    # Names are abbreviated in the workbook — kept as-is (not truncation, just short forms)
    "Emory": {
        "accept": 12, "satMedian": 1470, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Atlanta, GA",
        "vibe": "Research powerhouse in Atlanta; dominant pre-med culture and strong social scene",
    },
    "NYU": {
        "accept": 12, "satMedian": 1460, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "New York, NY",
        "vibe": "Urban campus without borders; Greenwich Village is your quad in the heart of NYC",
    },
    "Chicago": {
        "accept": 6, "satMedian": 1530, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Chicago, IL",
        "vibe": "Intellectual intensity above all else; famous for taking ideas more seriously than sleep",
    },
    "Washington (Mo)": {
        "accept": 14, "satMedian": 1500, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "St. Louis, MO",
        "vibe": "Research powerhouse in the Midwest; strong pre-med, engineering, and business",
    },
    "Carnegie Mellon": {
        "accept": 11, "satMedian": 1535, "hiddenIvy": False, "stem": True,
        "merit": "none", "location": "Pittsburgh, PA",
        "vibe": "Top CS and engineering with a rigorous, career-driven campus culture",
    },
    "Case Western": {
        "accept": 30, "satMedian": 1455, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Cleveland, OH",
        "vibe": "STEM-focused research university; pre-med and engineering define campus life",
    },
    "Rochester": {
        "accept": 29, "satMedian": 1440, "hiddenIvy": False, "stem": True,
        "merit": "high", "location": "Rochester, NY",
        "vibe": "Research-intensive with strong engineering, optics, and pre-med programs",
    },
    "Brandeis": {
        "accept": 37, "satMedian": 1420, "hiddenIvy": False, "stem": False,
        "merit": "moderate", "location": "Waltham, MA",
        "vibe": "Social justice mission and research strength near Boston with a unique founding story",
    },
    # ── MIAC ─────────────────────────────────────────────────────────────────
    "Gustavus Adolphus College": {
        "accept": 72, "satMedian": 1195, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Saint Peter, MN",
        "vibe": "Lutheran liberal arts with Swedish heritage and strong athletics in Minnesota",
    },
    "Carleton College": {
        "accept": 18, "satMedian": 1495, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Northfield, MN",
        "vibe": "One of the Midwest's best LACs; intellectual culture with top grad school placement",
    },
    "Saint John's University": {
        "accept": 75, "satMedian": 1145, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Collegeville, MN",
        "vibe": "Coordinate college with Saint Benedict; Benedictine tradition and strong community",
    },
    "Macalester College": {
        "accept": 28, "satMedian": 1430, "hiddenIvy": False, "stem": False,
        "merit": "none", "location": "Saint Paul, MN",
        "vibe": "Globally focused and politically active urban LAC with high international enrollment",
    },
    "Saint Olaf College": {
        "accept": 48, "satMedian": 1270, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Northfield, MN",
        "vibe": "Norwegian Lutheran roots; world-famous music programs and close Minnesota community",
    },
    "Hamline University": {
        "accept": 82, "satMedian": 1130, "hiddenIvy": False, "stem": False,
        "merit": "high", "location": "Saint Paul, MN",
        "vibe": "Urban liberal arts with social justice focus; personal and accessible in Twin Cities",
    },
}

# ---------------------------------------------------------------------------
# Data loading from Excel
# ---------------------------------------------------------------------------
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'data', 'lane4_swim_model.xlsx')

BENCHMARKS = {}    # "Conference|Event" -> {first, eighth, sixteenth, sec_per_place}
TEAMS = {}         # "Conference|School" -> {conference, school, psf, tier, finish, normalized}
TEAMS_LIST = []    # ordered list of all team dicts
CONFERENCES = {}   # conference name -> sorted list of canonical school names
NORMALIZATION_LOG = []  # records every name that was normalized

def load_data():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # ── Sheet1: BENCHMARKS ──────────────────────────────────────────────────
    ws = wb['Sheet1']
    h1 = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col1 = {v: i + 1 for i, v in enumerate(h1) if v}

    for r in range(2, ws.max_row + 1):
        conf  = ws.cell(r, col1['Conference']).value
        event = ws.cell(r, col1['Event']).value
        if not conf or not event:
            continue
        BENCHMARKS[f"{conf}|{event}"] = {
            'first':         _float(ws.cell(r, col1['1st_sec']).value),
            'eighth':        _float(ws.cell(r, col1['8th_sec']).value),
            'sixteenth':     _float(ws.cell(r, col1['16th_sec']).value),
            'sec_per_place': _float(ws.cell(r, col1['Sec_per_place']).value),
        }

    # ── Team_Tiers: TEAMS ────────────────────────────────────────────────────
    ws2 = wb['Team_Tiers']
    h2 = [ws2.cell(1, c).value for c in range(1, ws2.max_column + 1)]
    col2 = {}
    for i, v in enumerate(h2):
        if v and v not in col2:   # first-occurrence wins — sheet has duplicate headers at col 14
            col2[v] = i + 1

    for r in range(2, ws2.max_row + 1):
        conf   = ws2.cell(r, col2['Conference']).value
        raw    = ws2.cell(r, col2['Team']).value
        psf    = ws2.cell(r, col2['PSF']).value
        tier   = ws2.cell(r, col2['Tier']).value
        finish = ws2.cell(r, col2['Finish']).value
        pts    = ws2.cell(r, col2['MenPoints']).value
        if not conf or not raw:
            continue

        # Apply name normalization
        normalized = False
        canonical = raw
        if raw in TEAM_NAME_MAP:
            canonical, reason = TEAM_NAME_MAP[raw]
            normalized = True
            NORMALIZATION_LOG.append({
                'raw': raw, 'canonical': canonical,
                'reason': reason, 'conference': conf, 'finish': finish,
            })

        team_rec = {
            'conference':  conf,
            'school':      canonical,
            'raw_name':    raw,
            'psf':         _float(psf) if psf is not None else 1.0,
            'tier':        tier or '',
            'finish':      finish,
            'men_points':  _float(pts),
            'normalized':  normalized,
        }
        key = f"{conf}|{canonical}"
        TEAMS[key] = team_rec
        TEAMS_LIST.append(team_rec)

        if conf not in CONFERENCES:
            CONFERENCES[conf] = []
        if canonical not in CONFERENCES[conf]:
            CONFERENCES[conf].append(canonical)

    # Sort teams within each conference by finish position
    for conf in CONFERENCES:
        CONFERENCES[conf].sort(
            key=lambda s: TEAMS.get(f"{conf}|{s}", {}).get('finish') or 99
        )

def _float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

load_data()

# ---------------------------------------------------------------------------
# Scoring engine — all formulas from Swimmer_Calcs (workbook authoritative)
#
# Layer architecture (clean separation for later admissions layer):
#
#   SWIM LAYER  ──  _score_event()  →  EventScore
#                   _score_school_swim()  →  SwimResult
#
#   ADMISSION LAYER  ──  admission_chance(school, sat, gpa, adj_tier, psf)  →  AdmissionResult
#                        (takes SwimResult outputs + academic inputs; returns {label, color, …})
#
#   FULL PIPELINE  ──  score_all_schools(times, sat, gpa)  →  [SchoolResult …]
#                      (SwimResult + SCHOOL_META nested as `meta` + AdmissionResult)
#
# Output field names follow OUTPUT_SCHEMA.md exactly:
#   EventScore:  { event, sec, place, pts }   (+ expPts, confidence, placeLabel for tracing)
#   SchoolResult: { school, conference, tier, psf, rawPts, adjPts, adjTier,
#                   top3, allEvents, admission, meta, normalized, rawName }
# ---------------------------------------------------------------------------

ALL_EVENTS = [
    '50 Free', '100 Free', '200 Free', '500 Free',
    '1000 Free', '1650 Free',
    '100 Back', '200 Back',
    '100 Breast', '200 Breast',
    '100 Fly', '200 Fly',
    '200 IM', '400 IM',
    '50 Breast (Relay Split)',
]

# ── Primitives ─────────────────────────────────────────────────────────────

def parse_time(s):
    """'M:SS.ss' or 'SS.ss' → decimal seconds. Returns None if invalid/missing."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    try:
        if ':' in s:
            m, sec = s.split(':', 1)
            return float(m) * 60 + float(sec)
        return float(s)
    except ValueError:
        return None

def estimate_place(sec, b):
    """
    3-zone linear interpolation — workbook formula (Swimmer_Calcs):

      Zone 1  sec <= first      → 1.0          (capped; workbook IF behaviour)
      Zone 2  sec <= eighth     → 1 + (sec-1st)/(8th-1st) * 7
      Zone 3  sec <= sixteenth  → 8 + (sec-8th)/(16th-8th) * 8
      Zone 4  sec > sixteenth   → 16 + (sec-16th) / secPerPlace

    Returns a continuous float. No upper ceiling.
    """
    first     = b['first']
    eighth    = b['eighth']
    sixteenth = b['sixteenth']
    spp       = b['sec_per_place'] or 1.0

    if sec <= first:
        return 1.0
    if sec <= eighth:
        return 1.0 + (sec - first)  / ((eighth    - first)    or 1.0) * 7
    if sec <= sixteenth:
        return 8.0 + (sec - eighth) / ((sixteenth  - eighth)   or 1.0) * 8
    return 16.0 + (sec - sixteenth) / spp

def exp_points(place):
    """MAX(0, MIN(20, 21−place)) — workbook formula. Continuous float."""
    return max(0.0, min(20.0, 21.0 - place))

def confidence_weight(place):
    """
    Bubble-zone confidence discount — from Swimmer_Calcs.
    Full weight for A/B finalists; discounted for bubble; zero below 16th.
    """
    if place <= 12: return 1.0
    if place <= 14: return 0.85
    if place <= 16: return 0.65
    return 0.0

def place_label(place):
    """Human-readable place outcome — OUTPUT_SCHEMA thresholds."""
    if place <= 1.5:  return '🥇 Winner'
    if place <= 3.5:  return '🏅 Podium'
    if place <= 8.5:  return 'A Final'
    if place <= 16.5: return 'B Final'
    if place <= 20:   return 'Bubble'
    return 'Out of range'

def tier_label(pts):
    """
    Swim tier from adjPts — workbook thresholds (authoritative over spec).
    Called with rawPts for display and adjPts for the canonical tier.
    """
    if pts < 1:   return 'Moonshot'
    if pts < 4:   return 'Reach'
    if pts < 10:  return 'Recruitable'
    if pts < 18:  return 'Priority Recruit'
    if pts < 35:  return 'Likely Commit'
    if pts < 50:  return 'Conference Star'
    return 'High-Point Contender'

# ── Swim layer ──────────────────────────────────────────────────────────────

def _score_event(event, time_str, conf):
    """
    Score one event for one conference.
    Returns an EventScore dict or None if no benchmark / no valid time.

    EventScore (OUTPUT_SCHEMA):
      event      — event name
      sec        — time in decimal seconds
      place      — estimated decimal finish place
      pts        — confidence-weighted points (workbook: expPts × confidence)
                   NOTE: spec defines pts as integer from placeToPoints() lookup;
                   workbook uses continuous formula — workbook is authoritative.
      expPts     — points before confidence discount (for tracing)
      confidence — confidence multiplier applied (for tracing)
      placeLabel — human-readable outcome
    """
    sec = parse_time(time_str)
    if sec is None:
        return None

    bench = BENCHMARKS.get(f"{conf}|{event}")
    if bench is None:
        return None   # event not benchmarked in this conference

    place  = estimate_place(sec, bench)
    exp_pt = exp_points(place)
    cw     = confidence_weight(place)
    pts    = exp_pt * cw               # workbook weighted value → spec's `pts` field

    return {
        'event':      event,
        'sec':        round(sec, 3),
        'place':      round(place, 2),
        'pts':        round(pts, 2),   # OUTPUT_SCHEMA field name
        'expPts':     round(exp_pt, 2),
        'confidence': cw,
        'placeLabel': place_label(place),
    }

def _score_school_swim(team_rec, times):
    """
    Pure swim-value layer for one school.

    Input:  team_rec (from TEAMS_LIST), times dict from swimmer profile
    Output: SwimResult dict, or None if the swimmer has no scorable events here.

    SwimResult fields (OUTPUT_SCHEMA — swim-layer subset):
      school, conference, finish, tier, psf
      rawPts   — sum of top-3 pts values
      adjPts   — rawPts × psf
      adjTier  — tier label from adjPts
      top3     — up to 3 highest-scoring EventScore objects
      allEvents — all scored EventScore objects, sorted pts desc
      normalized, rawName — provenance flags
    """
    conf   = team_rec['conference']
    school = team_rec['school']
    psf    = team_rec['psf']

    all_events = []
    for event, time_str in times.items():
        es = _score_event(event, time_str, conf)
        if es is not None:
            all_events.append(es)

    # Sort by pts descending; top-3 drive rawPts
    all_events.sort(key=lambda e: e['pts'], reverse=True)
    top3    = all_events[:3]
    raw_pts = round(sum(e['pts'] for e in top3), 2)

    if raw_pts == 0:
        return None   # zero-score guardrail — school excluded from results

    adj_pts  = round(raw_pts * psf, 2)
    adj_tier = tier_label(adj_pts)

    return {
        'school':     school,
        'conference': conf,
        'finish':     team_rec['finish'],
        'tier':       team_rec['tier'],
        'psf':        psf,
        'rawPts':     raw_pts,
        'adjPts':     adj_pts,
        'adjTier':    adj_tier,
        'top3':       top3,
        'allEvents':  all_events,
        'normalized': team_rec['normalized'],
        'rawName':    team_rec['raw_name'],
    }

# ── Admission layer ─────────────────────────────────────────────────────────

def admission_chance(school, sat, gpa, adj_tier, psf):
    """
    Deterministic admission scoring — per LOGIC_RULES.md section 8.
    Consumes swim-layer outputs (adj_tier, psf) + academic profile (sat, gpa).

    Returns AdmissionResult:
      label      — one of 9 admission labels  (OUTPUT_SCHEMA)
      color      — hex color for UI
      total      — raw numeric score (debug)
      acadScore  — academic sub-score (debug)
      swimScore  — swim sub-score (debug)
    """
    meta = SCHOOL_META.get(school)

    if meta is None:
        return {'label': 'Unknown', 'color': '#94A3B8',
                'total': None, 'acadScore': None, 'swimScore': None}

    if meta.get('moonshot'):
        return {'label': 'Moonshot — Apply for Fun', 'color': '#6B7280',
                'total': None, 'acadScore': None, 'swimScore': None}

    sat_median = meta['satMedian']
    accept     = meta['accept']

    # Academic score
    sat_diff = sat - sat_median
    if sat_diff >= 80:    acad = 4
    elif sat_diff >= 30:  acad = 3
    elif sat_diff >= -30: acad = 2
    elif sat_diff >= -80: acad = 1
    else:                 acad = 0

    if gpa >= 3.9:     acad += 1   # GPA bonus
    if accept > 60:    acad += 1   # accessibility bonus

    # Swim score (inverse to program prestige — workbook PSF values: 0.70, 0.78, 1.00, 1.10, 1.20)
    if psf <= 0.78:   swim = 1    # elite — coach has least admissions leverage
    elif psf <= 0.85: swim = 2    # (placeholder; no school currently has psf=0.85)
    elif psf <= 1.00: swim = 3    # mid-tier
    else:             swim = 4    # weaker program — maximum coach leverage

    if adj_tier in ('High-Point Contender', 'Conference Star'):  swim += 2
    elif adj_tier in ('Likely Commit', 'Priority Recruit'):      swim += 1

    total = acad + swim

    if total >= 9:   label, color = 'Virtual Lock',        '#059669'
    elif total >= 8: label, color = 'Very Strong Chance',  '#10B981'
    elif total >= 7: label, color = 'Strong Chance',       '#34D399'
    elif total >= 6: label, color = 'Realistic Shot',      '#2563EB'
    elif total >= 5: label, color = 'Possible',            '#3B82F6'
    elif total >= 4: label, color = 'Reach with Support',  '#F59E0B'
    elif total >= 3: label, color = 'Major Reach',         '#EF4444'
    else:            label, color = 'Extreme Reach',       '#DC2626'

    return {'label': label, 'color': color,
            'total': total, 'acadScore': acad, 'swimScore': swim}

# ── Full pipeline ───────────────────────────────────────────────────────────

def score_all_schools(times, sat, gpa):
    """
    Score swimmer against all 76 programs.
    Returns list of SchoolResult dicts sorted by adjPts descending.

    Pipeline: swim layer → meta lookup → admission layer.
    Schools with rawPts == 0 (zero scorable events) are excluded entirely.

    SchoolResult (OUTPUT_SCHEMA):
      school, conference, tier, psf
      rawPts, adjPts, adjTier
      top3, allEvents
      admission  — AdmissionResult { label, color, total*, acadScore*, swimScore* }
      meta       — SchoolMeta nested object { accept, satMedian, hiddenIvy, stem,
                   merit, location, vibe, moonshot? }
      normalized, rawName  — provenance
    """
    results = []

    for team_rec in TEAMS_LIST:
        # ── Swim layer
        swim = _score_school_swim(team_rec, times)
        if swim is None:
            continue

        # ── Meta lookup (feeds admission layer and UI display)
        meta_raw = SCHOOL_META.get(swim['school'], {})
        meta = {
            'accept':    meta_raw.get('accept'),
            'satMedian': meta_raw.get('satMedian'),
            'hiddenIvy': meta_raw.get('hiddenIvy', False),
            'stem':      meta_raw.get('stem', False),
            'merit':     meta_raw.get('merit', ''),
            'location':  meta_raw.get('location', ''),
            'vibe':      meta_raw.get('vibe', ''),
        }
        if meta_raw.get('moonshot'):
            meta['moonshot'] = True

        # ── Admission layer (consumes swim outputs + academic profile)
        adm = admission_chance(swim['school'], sat, gpa, swim['adjTier'], swim['psf'])

        # ── Assemble SchoolResult (OUTPUT_SCHEMA shape)
        results.append({
            **swim,              # school, conference, finish, tier, psf,
                                 # rawPts, adjPts, adjTier, top3, allEvents,
                                 # normalized, rawName
            'admission': adm,    # { label, color, total, acadScore, swimScore }
            'meta':      meta,   # nested SchoolMeta object
        })

    results.sort(key=lambda r: r['adjPts'], reverse=True)
    return results

def score_one_school(times, conference, school):
    """
    Score arbitrary times at one specific school — for the manual calculator.
    Runs the same swim + admission pipeline as score_all_schools() for one school.
    """
    team_key  = f"{conference}|{school}"
    team_rec  = TEAMS.get(team_key)
    if team_rec is None:
        return {'error': f'School "{school}" not found in {conference}'}

    # Swim layer — also collect unscored events for display
    scored, unscored = [], []
    for event, time_str in times.items():
        sec = parse_time(time_str)
        if sec is None:
            continue
        es = _score_event(event, time_str, conference)
        if es is not None:
            scored.append({**es, 'time': time_str, 'benchmarked': True})
        else:
            unscored.append({'event': event, 'time': time_str,
                             'sec': sec, 'benchmarked': False})

    scored.sort(key=lambda e: e['pts'], reverse=True)
    top3    = scored[:3]
    raw_pts = round(sum(e['pts'] for e in top3), 2)
    psf     = team_rec['psf']
    adj_pts = round(raw_pts * psf, 2)
    adj_tier = tier_label(adj_pts)
    adm     = admission_chance(school, JAMES['sat'], JAMES['gpa'], adj_tier, psf)
    meta_raw = SCHOOL_META.get(school, {})

    return {
        'school':     school,
        'conference': conference,
        'tier':       team_rec['tier'],
        'psf':        psf,
        'rawPts':     raw_pts,
        'adjPts':     adj_pts,
        'adjTier':    adj_tier,
        'top3':       [e['event'] for e in top3],
        'events':     scored + unscored,
        'admission':  adm,
        'meta': {
            'accept':    meta_raw.get('accept'),
            'satMedian': meta_raw.get('satMedian'),
            'hiddenIvy': meta_raw.get('hiddenIvy', False),
            'stem':      meta_raw.get('stem', False),
            'merit':     meta_raw.get('merit', ''),
            'location':  meta_raw.get('location', ''),
            'vibe':      meta_raw.get('vibe', ''),
        },
        'normalized': team_rec['normalized'],
        'rawName':    team_rec['raw_name'],
    }

# ---------------------------------------------------------------------------
# Claude helpers
# ---------------------------------------------------------------------------

def _get_anthropic():
    """Return an Anthropic client, or None if no key is configured."""
    key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if not key:
        return None
    try:
        import anthropic
        return anthropic.Anthropic(api_key=key)
    except Exception:
        return None

def _pre_sort(results, query, eliminated, my_list):
    """
    Re-sort top-35 slice based on query intent — per LOGIC_RULES.md section 10.
    Excludes eliminated schools and my-list schools before sorting.
    Returns top 35 from the resulting list.
    """
    excl = set(eliminated) | set(my_list)
    pool = [r for r in results if r['school'] not in excl]

    q = query.lower()
    if any(k in q for k in ('prestig', 'best school', 'academic')):
        pool.sort(key=lambda r: (r['meta'].get('accept') or 999))
    elif any(k in q for k in ('stem', 'engineer', 'tech', 'med', 'science')):
        pool.sort(key=lambda r: (0 if r['meta'].get('stem') else 1))
    elif any(k in q for k in ('money', 'cost', 'afford', 'save')):
        rank = {'high': 0, 'moderate': 1, 'none': 2, '': 3}
        pool.sort(key=lambda r: rank.get(r['meta'].get('merit', ''), 3))
    elif any(k in q for k in ('star', 'podium', 'win', 'lead')):
        pool.sort(key=lambda r: -r['adjPts'])
    elif any(k in q for k in ('fun', 'social', 'happy', 'vibe')):
        pool.sort(key=lambda r: -(r['meta'].get('accept') or 0))
    elif 'hidden ivy' in q or 'ivy' in q:
        pool.sort(key=lambda r: (0 if r['meta'].get('hiddenIvy') else 1))
    # default: already sorted by adjPts desc from score_all_schools

    return pool[:35]

def _build_school_line(i, r):
    """Format one numbered line for the Claude search prompt."""
    vibe = (r['meta'].get('vibe') or '')[:60]
    return (
        f"{i+1}. {r['school']} ({r['conference']}): "
        f"swimTier={r['adjTier']}, admission={r['admission']['label']}, "
        f"hiddenIvy={str(r['meta'].get('hiddenIvy', False)).lower()}, "
        f"stem={str(r['meta'].get('stem', False)).lower()}, "
        f"merit={r['meta'].get('merit', '')}, "
        f"accept={r['meta'].get('accept', '?')}%, "
        f"vibe=\"{vibe}\""
    )

def _parse_search_response(text, sorted_35):
    """
    Parse Claude's JSON search response.
    Returns list of enriched SchoolResult dicts (with aiWhy), or raises ValueError.
    """
    # Strip markdown fences
    text = re.sub(r'```json\s*', '', text)
    text = re.sub(r'```\s*', '', text)
    text = text.strip()

    match = re.search(r'\{[\s\S]*\}', text)
    if not match:
        raise ValueError('No JSON object found in response')

    parsed = json.loads(match.group())
    answer  = parsed.get('answer', '')
    picks   = parsed.get('schools', [])

    schools = []
    for pick in picks:
        idx = pick.get('number')
        if idx is None:
            continue
        idx = int(idx) - 1
        if idx < 0 or idx >= len(sorted_35):
            continue
        r = dict(sorted_35[idx])
        r['aiWhy'] = pick.get('why', '')
        schools.append(r)

    if not schools:
        raise ValueError('No valid school picks in response')

    return answer, schools

def _build_top3_text(top3):
    """'1650 Free: 🥇 Winner; 500 Free: 🏅 Podium' style string."""
    return '; '.join(f"{e['event']}: {place_label(e['place'])}" for e in top3)

def _build_vibe_lines(vibe):
    """Format answered vibe questions for deep dive prompt."""
    if not vibe:
        return ''
    labels = {
        'campus':   'Ideal campus feel',
        'friday':   'Friday night preference',
        'academic': 'Academic priority',
        'compete':  'Competition mindset',
        'location': 'Location preference',
        'career':   'Career interest',
    }
    lines = []
    for k, v in vibe.items():
        if v:
            lines.append(f"  - {labels.get(k, k)}: {v}")
    return '\n'.join(lines)

# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return send_from_directory('static', 'index.html')

@app.route('/api/meta', methods=['GET'])
def meta():
    return jsonify({
        'conferences':       sorted(CONFERENCES.keys()),
        'teams':             CONFERENCES,
        'events':            ALL_EVENTS,
        'normalizationLog':  NORMALIZATION_LOG,
    })

@app.route('/api/score-all', methods=['GET'])
def score_all():
    """Score James against all 76 programs. No request body needed."""
    results = score_all_schools(JAMES['times'], JAMES['sat'], JAMES['gpa'])
    return jsonify({
        'profile': JAMES,
        'totalSchools': len(TEAMS_LIST),
        'scoredSchools': len(results),
        'results': results,
    })

@app.route('/api/search', methods=['POST'])
def search():
    """
    Natural-language search: re-sort top-35 by query intent, call Claude,
    return 3 annotated SchoolResult objects.

    Body: { query, eliminated?, myList? }
    Response: { answer, schools } or { error, fallback? }
    """
    data       = request.json or {}
    query      = data.get('query', '').strip()
    eliminated = data.get('eliminated', [])
    my_list    = data.get('myList', [])

    if not query:
        return jsonify({'error': 'Query is required'}), 400

    # Always compute fresh so filtering/sorting is accurate
    all_results = score_all_schools(JAMES['times'], JAMES['sat'], JAMES['gpa'])
    sorted_35   = _pre_sort(all_results, query, eliminated, my_list)

    if not sorted_35:
        return jsonify({'error': 'No schools available to search after filters'}), 400

    client = _get_anthropic()
    if not client:
        return jsonify({
            'error': 'AI search is not configured',
            'detail': 'ANTHROPIC_API_KEY is missing or invalid',
            'fallback': sorted_35[:3],  # return top 3 deterministically
        }), 503

    system_prompt = (
        "You are Lane4. Respond ONLY with a valid JSON object. "
        "No markdown. No explanation. Start with { end with }. "
        "Keep 'why' fields under 15 words each. Keep 'answer' under 30 words."
    )

    school_lines = '\n'.join(_build_school_line(i, r) for i, r in enumerate(sorted_35))
    user_prompt  = (
        f'Mom\'s question: "{query}"\n\n'
        f"James: GPA {JAMES['gpa']}, SAT {JAMES['sat']}, STEM-focused, "
        f"Conference Star swimmer (1650/500 Free, 50 Breast split).\n\n"
        "Pick 3 schools from this numbered list that best answer the question. Return ONLY JSON.\n\n"
        f"{school_lines}\n\n"
        'JSON format:\n{"answer":"1-2 sentences max","schools":[{"number":1,"why":"under 15 words"}]}'
    )

    try:
        resp = client.messages.create(
            model='claude-sonnet-4-5-20250929',
            max_tokens=600,
            system=system_prompt,
            messages=[{'role': 'user', 'content': user_prompt}],
        )
        raw_text = resp.content[0].text
        answer, schools = _parse_search_response(raw_text, sorted_35)
        return jsonify({'answer': answer, 'schools': schools})

    except json.JSONDecodeError as e:
        return jsonify({
            'error': 'AI returned malformed JSON',
            'detail': str(e),
            'fallback': sorted_35[:3],
        }), 200

    except ValueError as e:
        return jsonify({
            'error': str(e),
            'fallback': sorted_35[:3],
        }), 200

    except Exception as e:
        return jsonify({
            'error': 'Search failed',
            'detail': str(e),
            'fallback': sorted_35[:3],
        }), 200


@app.route('/api/deep-dive', methods=['POST'])
def deep_dive():
    """
    Generate the 8-section deep dive narrative for one school.

    Body: { school }   (school must match a key in score_all results)
    Response: { sections: [{title, body}] } or { error }
    """
    data    = request.json or {}
    school  = data.get('school', '').strip()

    if not school:
        return jsonify({'error': 'school is required'}), 400

    all_results = score_all_schools(JAMES['times'], JAMES['sat'], JAMES['gpa'])
    result = next((r for r in all_results if r['school'] == school), None)

    if result is None:
        return jsonify({'error': f'School "{school}" not found in scored results'}), 404

    client = _get_anthropic()
    if not client:
        return jsonify({
            'error': 'AI deep dive is not configured',
            'detail': 'ANTHROPIC_API_KEY is missing or invalid',
        }), 503

    meta = result['meta']
    top3_text = _build_top3_text(result['top3'])
    vibe_lines = _build_vibe_lines(JAMES.get('vibe'))

    merit_label = {
        'none':     'Need-based only',
        'high':     'Strong merit aid available',
        'moderate': 'Moderate merit aid',
    }.get(meta.get('merit', ''), 'Moderate merit aid')

    vibe_block = ''
    if vibe_lines:
        vibe_block = (
            f"\n{JAMES['name']}'S PERSONALITY & PREFERENCES "
            f"(use these to personalize every section):\n{vibe_lines}\n"
        )

    hidden_ivy_note = '\nThis is a Hidden Ivy — academically elite, employer-respected, without the brand tax.' if meta.get('hiddenIvy') else ''
    stem_note       = '\nStrong STEM programs.' if meta.get('stem') else ''

    system_prompt = (
        "You are Lane4, a college swim recruiting advisor. "
        "Warm, honest, direct. Talk to a 17-year-old and his mom. "
        "Never use jargon. 'Hidden Ivy' means academically elite and employer-respected "
        "without the Stanford rejection rate. The comp anchor — comparing to a dream school "
        "— is powerful when honest."
    )

    user_prompt = (
        f"Write a deep dive for {JAMES['name']} considering {result['school']}.\n\n"
        f"SWIMMER: {JAMES['name']}, Class of 2026, GPA {JAMES['gpa']} unweighted, "
        f"SAT {JAMES['sat']} (math {JAMES['mathSat']}), "
        f"projected retake {JAMES['satProjected']} (math {JAMES['mathSatProjected']}). "
        f"STEM-focused. Heavy AP load including Calc BC."
        f"{vibe_block}\n"
        f"SWIM RESULTS AT {result['school'].upper()} ({result['conference']}):\n"
        f"Top events: {top3_text}\n"
        f"Conference tier (raw): {tier_label(result['rawPts'])}\n"
        f"Program adjusted tier (PSF {result['psf']}): {result['adjTier']}\n"
        f"Admission outlook: {result['admission']['label']}"
        f"{hidden_ivy_note}{stem_note}\n"
        f"School vibe: {meta.get('vibe', '')}\n"
        f"Location: {meta.get('location', '')}\n"
        f"Acceptance rate: ~{meta.get('accept', '?')}%\n"
        f"SAT median: ~{meta.get('satMedian', '?')}\n"
        f"Merit aid: {merit_label}\n\n"
        "Write exactly these sections. Warm, direct, honest. Talk to a 17-year-old and his mom. "
        "Never clinical. Weave in what you know about his personality — don't just list his "
        "preferences, speak to them naturally. Use 'Hidden Ivy' naturally if applicable. "
        "Max 2-3 sentences per section.\n\n"
        "## Your Honest Shot\n"
        "## What This School Is Actually Like\n"
        f"## How {JAMES['name']} Fits on the Swim Team\n"
        "## Why a Coach Would Want to Call\n"
        "## Getting In — The Real Picture\n"
        "## The Money Conversation\n"
        "Include: Estimated COA, Estimated Merit Aid for this profile, Estimated Net Cost\n"
        "## Your Next Three Moves\n"
        "Three specific actions this week.\n"
        "## The Bottom Line\n"
        "One sentence. Make it land."
    )

    try:
        resp = client.messages.create(
            model='claude-sonnet-4-5-20250929',
            max_tokens=1000,
            system=system_prompt,
            messages=[{'role': 'user', 'content': user_prompt}],
        )
        raw = resp.content[0].text

        # Split on section headers — per OUTPUT_SCHEMA response parsing spec
        parts  = re.split(r'^## ', raw, flags=re.MULTILINE)
        sections = []
        for part in parts:
            part = part.strip()
            if not part:
                continue
            lines = part.split('\n', 1)
            title = lines[0].strip()
            body  = lines[1].strip() if len(lines) > 1 else ''
            if title:
                sections.append({'title': title, 'body': body})

        if not sections:
            return jsonify({'error': 'AI returned empty deep dive', 'raw': raw}), 200

        return jsonify({
            'school':    result['school'],
            'sections':  sections,
            'admission': result['admission'],
            'adjTier':   result['adjTier'],
            'meta':      meta,
        })

    except Exception as e:
        return jsonify({'error': 'Deep dive failed', 'detail': str(e)}), 200


@app.route('/api/coach-email', methods=['POST'])
def coach_email():
    """
    Generate deterministic coach email for one school. No AI call.

    Body: { school }
    Response: { subject, body }
    """
    data   = request.json or {}
    school = data.get('school', '').strip()

    if not school:
        return jsonify({'error': 'school is required'}), 400

    all_results = score_all_schools(JAMES['times'], JAMES['sat'], JAMES['gpa'])
    result = next((r for r in all_results if r['school'] == school), None)

    if result is None:
        return jsonify({'error': f'School "{school}" not found in scored results'}), 404

    meta  = result['meta']
    top3  = result['top3']
    best  = top3[0] if top3 else None

    if best is None:
        return jsonify({'error': 'No scored events found for this school'}), 400

    # Performance descriptor
    if best['place'] <= 1.5:
        perf = f"projected to win the {best['event']}"
    elif best['place'] <= 3.5:
        perf = f"projected to podium in the {best['event']}"
    else:
        perf = f"projected as a conference A finalist in the {best['event']}"

    second   = f" I also project to score in the {top3[1]['event']}." if len(top3) > 1 else ''
    stem_note  = ' Your programs in engineering and CS align directly with my academic direction.' if meta.get('stem') else ''
    merit_note = " I've also been looking closely at your merit scholarship opportunities." if meta.get('merit') == 'high' else ''

    subject = f"Prospective Student-Athlete Inquiry — Class of 2026 | Distance Freestyle"
    body = (
        f"Dear Coach,\n\n"
        f"My name is {JAMES['name']} and I'm a junior in the Class of 2026 with strong interest "
        f"in {result['school']}'s swim program.\n\n"
        f"At the {result['conference']} conference level, I'm {perf}.{second} "
        f"My current bests include a 16:06 in the 1650 free, 4:37 in the 500, "
        f"and a 25.68 relay split in the 50 breast.\n\n"
        f"Academically I carry a {JAMES['gpa']} GPA with a heavy AP load including Calc BC "
        f"and scored a {JAMES['sat']} SAT with a retake planned.{stem_note}{merit_note}\n\n"
        f"I'd love to connect about your program. Would you have time for a brief call or campus visit?\n\n"
        f"Thank you,\n{JAMES['name']}"
    )

    return jsonify({'subject': subject, 'body': body})


@app.route('/api/health', methods=['GET'])
def health():
    key_ok = bool(os.environ.get('ANTHROPIC_API_KEY', '').strip())
    return jsonify({
        'status':        'ok',
        'benchmarks':    len(BENCHMARKS),
        'teams':         len(TEAMS_LIST),
        'schoolMeta':    len(SCHOOL_META),
        'normalized':    len(NORMALIZATION_LOG),
        'anthropicKey':  key_ok,
    })

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
